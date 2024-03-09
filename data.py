from openpyxl.reader.excel import load_workbook

class WebData:

    """
    This class is used to contain all the data that are required to perform testing for the OrangeHRM
    The test data are given from the HRMExcel.xlsx
    """
    def __init__(self):
        self.url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
        self.dashboardURL = "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index"
        self.fileName = "Data/HRMExcel.xlsx"
        self.sheetName = "Sheet1"
        self.workbook = load_workbook(self.fileName)
        self.sheet = self.workbook[self.sheetName]

    def row_count(self):
        return self.sheet.max_row

    def read_data(self, row, column):
        """
         This method will return data present in a particular cell in the sheet1
        :return
        """
        return self.sheet.cell(row, column).value

    def write_data(self, row, column, data):
        """
        This method is used to assign the data to the particular cell in the sheet
        :param row:
        :param column:
        :return:
        """
        self.sheet.cell(row, column).value = data
        self.workbook.save(self.fileName)

